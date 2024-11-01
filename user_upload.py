import os
import pandas as pd
import pickle
import streamlit as st
import re

import uploadStruc
import uploadStrucZipClass
import uploadStrucbyZIP
import utils

from quickstart_googledrive import upload_or_replace_file, st_upload_file_to_drive, file_from_gdrive, dir_dict
from quickstart_googledrive import init_drive_client

from ReadExcel_df import collect_errors_and_styles
from columns import column_title_dict, adsorbate_dic
from utils import get_total_excel, is_valid_doi

from user import login
from datetime import datetime
import datetime as mydt
import time

from openpyxl import load_workbook

# 反应名称及简写
sheet_dict = {
    "Oxygen Reduction Reaction": "ORR",  # 1
    "Oxygen Evolution Reaction": "OER",  # 2
    "Hydrogen Evolution Reaction": "HER",  # 3
    "Carbon Dioxide Reduction": "CO2RR",  # 4
    "Carbon Monoxide Reduction": "CORR",
    "Ammonia Synthesis": "NRR",  # 5
    "Hydrogen Oxidation Reaction": "HOR",  # 6
    "Hydrogen Peroxide Synthesis": "water oxidation-H2O2",  # 7
    "Epoxide Production": "Epoxide_Production",  # 8
    "Chlorination Evolution Reaction": "Chlorination evolution reaction",  # 9
    "Ozone Evolution Reaction": "O3ER",  # 10
    "Nitrogen Oxidation Reaction": "NOR",  # 11
    "NH3 Oxidation Reaction": "NH3OR",  # 12
    "Urea Oxidation Reaction": "UOR",  # 13
    "Hydrazine Oxidation Reaction": "HzOR",  # 14
    "Urea Synthesis": "Urea Synthesis",  # 15
    "Nitrate Reduction": "Nitrate Reduction",  # 16
    "Electrocatalytic hydrogenation": "Electrocatalytic hydrogenation",  # 17
    "CH4 Oxidation Reaction": "CH4OR",  # 18
    "Methanol Ethanol Reforming": "Methanol Ethanol Reforming",  # 19
    "Organic Electrocatalysis": "Organic Electrocatalysis",  # 20
    "CO Oxidation Reaction": "CO Oxidation Reaction",  # 21
    "5-hydroxymethylfurfural (HMF) Oxidation": "HMF oxidation",  # 22
    "Acetylene semihydrogenation": "Acetylene semihydrogenation",  # 23
    "Ammoximation Reaction": "Ammoximation Reaction",  # 24
    "Oxygen-containing radical synthesis": "Oxygen-containing radical",

    # Thermocatalysis
    "Propane dehydrogenation": "Propane dehydrogenation",  # 1
    "Carbon monoxide oxidation": "Carbon monoxide oxidation",
    "Carbon dioxide hydrogenation": "CO2hydro",
    "Thermal Ammonia Synthesis": "Ammonia",
    "Carbon dioxide + methonal = DMC(dimethyl carbonate)": "CO2methonal2DMC",
    "Fischer-Tropsch Synthesis": "Fischer Tropsch Synthesis",
    "Methanol Oxidative Carbonylation to DMC": "Methanol Oxidative to DMC",

    # Photocatalysis
    "Solar Water Splitting": "Solar Water Splitting",
    "HMF oxidation": "HMF oxidation",
    "Oxygen-containing radical": "Oxygen-containing radical",
    "CO2methonal2DMC": "CO2methonal2DMC",

}


def is_string_in_file(file_path, string):
    try:
        with open(file_path, 'r') as file:
            content = file.read()
            if string in content:
                return True
            else:
                return False
    except FileNotFoundError:
        return False


def read_excel_data(file_path, sheet_name=0, row=None, column=None):
    try:
        # Attempt to read the specified sheet
        df = pd.read_excel(file_path, sheet_name=sheet_dict[sheet_name])
        if row is not None and column is not None:
            return df.iloc[row, column]
        elif row is not None:
            return df.iloc[row].tolist()
        elif column is not None:
            return df[column].tolist()
        else:
            return df
    except ValueError:
        # If the sheet does not exist, return an empty DataFrame
        df = pd.DataFrame()
        return df


def streamlit_frame():
    st.set_page_config(
        page_title="Contribution to Digcat",
        page_icon="🎏",
        layout='wide'
    )

    # st.sidebar.image("logo.svg")
    st.sidebar.markdown('<h1 style="font-size: 2em;">Contribution to Digcat</h1>', unsafe_allow_html=True)
    st.sidebar.markdown("# 	:roller_coaster:Entry:")
    st.image("title.svg", use_column_width=True)

    thermo_on = st.sidebar.toggle("Switching to Thermocatalysis", )
    photo_on = st.sidebar.toggle("Switching to Photocatalysis", disabled=thermo_on)

    st.sidebar.header("Please choose the reaction type:")

    # get reaction name
    st.sidebar.markdown("### :robot_face: Electrocatalysis:")

    reaction_name = st.sidebar.selectbox(" ", (
        "1. Oxygen Reduction Reaction",
        "2. Oxygen Evolution Reaction",
        "3. Hydrogen Evolution Reaction",
        "4. Carbon Dioxide Reduction",
        "5. Carbon Monoxide Reduction",
        "6. Ammonia Synthesis",
        "7. Hydrogen Oxidation Reaction",
        "8. Hydrogen Peroxide Synthesis",
        "9. Epoxide Production",
        "10. Chlorination Evolution Reaction",
        "11. Ozone Evolution Reaction",
        "12. Nitrogen Oxidation Reaction",
        "13. NH3 Oxidation Reaction",
        "14. Urea Oxidation Reaction",
        "15. Hydrazine Oxidation Reaction",
        "16. Urea Synthesis",
        "17. Nitrate Reduction",
        "18. Electrocatalytic hydrogenation",
        "19. CH4 Oxidation Reaction",
        "20. Methanol Ethanol Reforming",
        "21. Organic Electrocatalysis",
        "22. CO Oxidation Reaction",
        "23. 5-hydroxymethylfurfural (HMF) Oxidation",
        "24. Acetylene semihydrogenation",
        "25. Ammoximation Reaction",
        "26. Oxygen-containing radical synthesis",
        "27. Magnetic electrocatalysis",
        "28. Seawater electrolysis"
    ), disabled=st.session_state["modify_data"] or thermo_on or photo_on, label_visibility="collapsed")

    if reaction_name == "27. Magnetic electrocatalysis":
        reaction_name_mag = st.sidebar.selectbox(" ", (
            "1. Oxygen Reduction Reaction",
            "2. Oxygen Evolution Reaction",
            "3. Hydrogen Evolution Reaction",
            "4. Carbon Dioxide Reduction",
            "5. Carbon Monoxide Reduction"
            "6. Ammonia Synthesis",
            "7. Hydrogen Oxidation Reaction",
            "8. Hydrogen Peroxide Synthesis",
            "9. Epoxide Production",
            "10. Chlorination Evolution Reaction",
            "11. Ozone Evolution Reaction",
            "12. Nitrogen Oxidation Reaction",
            "13. NH3 Oxidation Reaction",
            "14. Urea Oxidation Reaction",
            "15. Hydrazine Oxidation Reaction",
            "16. Urea Synthesis",
            "17. Nitrate Reduction",
            "18. Electrocatalytic hydrogenation",
            "19. CH4 Oxidation Reaction",
            "20. Methanol Ethanol Reforming",
            "21. Organic Electrocatalysis",
            "22. CO Oxidation Reaction",
            "23. 5-hydroxymethylfurfural (HMF) Oxidation",
            "24. Acetylene semihydrogenation",
            "25. Ammoximation Reaction",
            "26. Oxygen-containing radical synthesis",
        ), disabled=st.session_state["modify_data"] or thermo_on or photo_on, label_visibility="collapsed")

    if reaction_name == "28. Seawater electrolysis":
        reaction_name_sea = st.sidebar.selectbox(" ", (
            "1. Oxygen Evolution Reaction",
            "2. Hydrogen Evolution Reaction",
            "3. Chlorination Evolution Reaction",
        ), disabled=st.session_state["modify_data"] or thermo_on or photo_on, label_visibility="collapsed")

    st.sidebar.markdown("\n\n\n\n\n\n\n\n\n\n")

    st.sidebar.markdown("### :hotsprings: Thermocatalysis:")
    thermo_reaction_name = st.sidebar.selectbox(" ", (
        "1. Propane dehydrogenation",
        "2. Carbon monoxide oxidation",
        "3. Carbon dioxide hydrogenation",
        "4. Thermal Ammonia Synthesis",
        "5. Carbon dioxide + methonal = DMC(dimethyl carbonate)",
        "6. Fischer-Tropsch Synthesis",
        "7. Methanol Oxidative Carbonylation to DMC"
    ), disabled=st.session_state["modify_data"] or not thermo_on, label_visibility="collapsed")

    st.sidebar.markdown("\n\n\n\n\n\n\n\n\n\n")
    st.sidebar.markdown("### :sunny: Photocatalysis:")
    photo_reaction_name = st.sidebar.selectbox(" ", (
        "1. Solar Water Splitting",
    ), disabled=st.session_state["modify_data"] or not photo_on or thermo_on, label_visibility="collapsed")

    if thermo_on:
        reaction_name = thermo_reaction_name.split(".")[-1].strip()
    elif photo_on:
        reaction_name = photo_reaction_name.split(".")[-1].strip()
    else:
        reaction_name = reaction_name.split(".")[-1].strip()
    mag_flag = False
    sea_flag = False
    if reaction_name == "Magnetic electrocatalysis":
        reaction_name = reaction_name_mag.split(".")[-1].strip()
        mag_flag = True
    elif reaction_name == "Seawater electrolysis":
        reaction_name = reaction_name_sea.split(".")[-1].strip()
        sea_flag = True
    if not st.session_state["modify_data"]:
        st.session_state["reaction_name"] = reaction_name

    st.sidebar.write(f"You selected {st.session_state['reaction_name']}.")
    if thermo_on:
        photo_on = False
        st.sidebar.write("You selected Thermocatalysis.")
    elif photo_on:
        thermo_on = False
        st.sidebar.write("You selected photocatalysis.")
    else:
        st.sidebar.write("You selected electrocatalysis.")

    return st.session_state["reaction_name"], mag_flag, thermo_on, sea_flag, photo_on


def reaction_column_title_dict(reaction_name, mag_flag, sea_flag):
    _column_title_dict = {
        "Oxygen Reduction Reaction":
            {
                "Formula": "e.g. PtNi",
                "Type": "",
                "Elements": "e.g. Pt,Ni",
                "Content": "e.g. Pt/Ni=60.7/39.3",
                "Structural Property": "e.g. [PtNi(111)] or [Co site on Co-porphyrin-N surface]",
                "Electrolytes": "e.g. 0.1 M HClO4",
                "pH": "e.g. 0~14",
                "Rotation_Speed (rpm)": "e.g. 1600",
                "Half-wave Potential (V/RHE)": "e.g. 0.90",
                "On-set potential, disk electrode (V/RHE@j mA cm-2)": "e.g. 0.95@0.5",
                "On-set potential, ring electrode (V/RHE@j mA cm-2)": "e.g. 0.65@0.1",
                "Disk total current density (mAcm-2@V/RHE)": "e.g. 3.0@0.9",
                "Disk total TOF (s-1@V/RHE)": "e.g. 1.0@0.9",
                "Disk ring current density (mAcm-2@V/RHE)": "e.g. 0.1@0.6",
                "Disk ring TOF (s-1@V/RHE)": "e.g. 1.0@0.6",
                "H2O2 faradaic Efficiency (%@V/RHE)": "e.g. 80%@0.6",
                "ECSA(HUPD (m2 g-1)": "e.g. 43.5",
                "Mass Activity (A mg-1@V/RHE)": "e.g. 1.89@0.9",
                "Specific Acticity (mA cm-2@V/RHE)": "e.g. 4.34@0.9",
                "Stability Test (h@V/RHE)": "e.g. 30000@0.6",
                "Durability Loss %(Performace metrics)": "e.g. 13.7%(MA)",
                "Structural change after tests": "Describe the structure or content change after durability tests",
                "Year": "e.g. 2019"
            },
        "Oxygen Evolution Reaction":
            {
                "Formula": "e.g. RuCu",
                "Type": "",
                "Elements": "e.g. Ru,Cu",
                "Content": "Ru/Cu=5.5/1",
                "Structural Property": "e.g. Nanosheet-Ru(100)",
                "Catalyst loading on electrode (mg cm-2)": "e.g. 1.1",
                "Working electrode": "e.g. glassy carbon electrode",
                "Electrolytes": "e.g. 1 M KOH",
                "pH": "e.g. 13.6",
                "Over_Potential (mV@ j mA cm-2)": "e.g. 234@10",
                "Current_density (mA/cm2@ U V/RHE)": "e.g. 40@1.79",
                "Tafel_slope (mVdev-1)": "e.g. 80",
                "Device test": "Please enter yes/no",
                "Stability Test (h@ mAcm-2)": "e.g. 45@5",
                "Durability Loss %(performace metrics)": "e.g. 5%(Current_density)",
                "Year": "e.g. 2019",
            },
        "Hydrogen Evolution Reaction":
            {
                "Formula": "e.g. RuCu",
                "Type": "",
                "Elements": "e.g. Ru,Cu",
                "Content": "Ru/Cu=5.5/1",
                "Structural Property": "e.g. Nanosheet-Ru(100)",
                "Catalyst loading on electrode (mg cm-2)": "e.g. 1.1",
                "Working electrode": "e.g. glassy carbon electrode",
                "Electrolytes": "e.g. 1 M KOH",
                "Over_Potential (mV@ j mA cm-2)": "e.g. 20@10",
                "Tafel_slope (mVdev-1)": "e.g. 15.3",
                "Device test": "Please enter yes/no",
                "Stability Test (h@ mAcm-2)": "e.g. 45@5",
                "Durability Loss %(performace metrics)": "e.g. 5%(Current_density)",
                "Year": "e.g. 2019",
            },
        "Carbon Dioxide Reduction":
            {
                "Formula": "e.g. Ni-Zn-N-C",
                "Type": "",
                "Elements": "e.g. Ni,Zn",
                "Content": "e.g. Ni/Zn=0.84/1.32",
                "Structural Property": "e.g. Zn-Ni-N6-C, Dual-atom M-N-C catalysts",
                "Electrolytes": "0.5 M KHCO3",
                "Temperature (Celsius)": "e.g. 60",
                "pressure (bar)": "e.g. 1",
                "Type of cell": "e.g. AEM flow cell",
                "pH": "e.g. 6.8",
                "Main product": "",
                "Other product (FE > 10%)": "",
                "Main product faradaic efficiency (%@U V/RHE)": "e.g. 99%@-0.73",
                "Current density (mA cm-2@U V/RHE)": "e.g. 17.5@-0.8",
                "Stability Test (h@ mAcm-2)": "e.g. 28@14",
                "Durability Loss %(performace metrics)": "e.g. 4%(faradaic efficiency)",
                "Year": "e.g. 2021",
            },
        "Carbon Monoxide Reduction":
            {
                "Formula": "e.g. Ni-Zn-N-C",
                "Type": "",
                "Elements": "e.g. Ni,Zn",
                "Content": "e.g. Ni/Zn=0.84/1.32",
                "Structural Property": "e.g. Zn-Ni-N6-C, Dual-atom M-N-C catalysts",
                "Electrolytes": "0.5 M KHCO3",
                "Temperature (Celsius)": "e.g. 60",
                "pressure (bar)": "e.g. 1",
                "Type of cell": "e.g. AEM flow cell",
                "pH": "e.g. 6.8",
                "Main product": "",
                "Other product (FE > 10%)": "",
                "Main product faradaic efficiency (%@U V/RHE)": "e.g. 99%@-0.73",
                "Current density (mA cm-2@U V/RHE)": "e.g. 17.5@-0.8",
                "Stability Test (h@ mAcm-2)": "e.g. 28@14",
                "Durability Loss %(performace metrics)": "e.g. 4%(faradaic efficiency)",
                "Year": "e.g. 2021",
            },
        "Ammonia Synthesis":
            {
                "Formula": "e.g. Au/Ni",
                "Type": "",
                "Elements": "e.g. Au,Ni",
                "Content": "e.g. Au/Ni=6/1",
                "Structural Property": "e.g. Adjacent Au NPs and Ni NPs on Nitrogen-doped carbon",
                "Isotope labeling": "",
                "Electrolytes": "e.g. N2-saturated 0.05 M H2SO4",
                "pH": "e.g. 1.0",
                "Reactant": "",
                "Quantification method": "e.g. spectrophotometer or NMR",
                "faradaic Efficiency (%@U V/RHE)": "67.8%@-0.14",
                "Current density (mA cm-2@U V/RHE)": "e.g. 17.5@-0.8",
                "Rate (μg h-1 mg-1@U V/RHE)": "7.4@-0.14",
                "Year": "e.g. 2019",
            },
        "Hydrogen Oxidation Reaction":
            {
                "Formula": "e.g. 3%Pt-Ru/C",
                "Type": "",
                "Elements": "e.g. Pt,Ru",
                "Content": "e.g. Pt/Ru=3/97",
                "Structural Property": "e.g. Pt single atom in Ru-N-C structures",
                "Electrolytes": "e.g. 0.1 M KOH",
                "pH": "e.g. 13",
                "Exchange current densities (mA cm-2 @ V/RHE)": "e.g. 4.16@0.05",
                "Mass-normalized kinetic current density (mA cm-2 ug-1@U mV/RHE)": "1.71@50",
                "Mass activity (A/mg @ V/RHE)": "e.g. 0.65@0.05",
                "Stability Test (h@ mAcm-2)": "e.g. 100@1.71",
                "Year": "e.g. 2024",
            },
        "Hydrogen Peroxide Synthesis":
            {
                "Formula": "e.g. LaAlO3/FTO",
                "Type": "",
                "Elements": "e.g. La,Al",
                "Content": "e.g. La/Al=9.6/9.1",
                "Structural Property": "e.g. LaAlO3 (100）and (110）",
                "Working electrode": "e.g. LaAlO3/FTO",
                "Counter electrode": "e.g. Carbon rod",
                "Reference electrode": "e.g. Ag/AgCl (KCl Sat.)",
                "Electrolytes": "e.g. 4M K2CO3/ KHCO3",
                "pH": "e.g. 8.3",
                "Rotation_Speed (rpm)": "e.g. 1000",
                "Over_Potential(mV@j mAcm-2)": "e.g. 510@10",
                "On-set potential,three-electrode (V/RHE@j mA cm-2 )": "2.1@2",
                "Total current density (mAcm-2@U V/RHE)": "e.g. 18.5@2.4",
                "H2O2 current density (mAcm-2@U V/RHE)": "e.g. 34.2@3.2",
                "H2O2 Measurement Method": "e.g. UV-vis and  titration",
                "H2O2_peak faradaic_Efficiency (%@ U/RHE)": "e.g. 87@3.34",
                "Maximum  H2O2 production rates（mmol cm-2 min-1)/ppm min-1 @ V/RHE": "e.g. 0.00089@2.7",
                "ECSA(HUPD)(m2 g-1)": "e.g. 2.76",
                "Stability Test (h@U V/RHE)": "e.g. 3@2.7",
                "Durability Loss %(performace metrics)": "e.g. 3%(faradaic efficiency)",
                "Year": "e.g. 2022",
            },
        "Epoxide Production":
            {
                "Formula": "e.g. Pd3Pt1",
                "Type": "",
                "Elements": "e.g. Pd,Pt",
                "Content": "e.g. Pd/Pt=3/1",
                "Structural Property": "e.g. Pd3Pt1 (111)",
                "Catalyst loading on electrode (mg cm-2)": "e.g. 0.5",
                "Electrolytes": "e.g. 0.4 M TBABF4",
                "pH": "e.g. 6",
                "Halogen-Mediated": "yes/no",
                "Current density (mAcm-2@U V/RHE)": "e.g. 18.5@2.4",
                "Highest epoxide production rate (μmol/cm2/h)": "9.2",
                "Epoxidation FE at the highest rate (%@V/RHE)": "61%",
                "Main product": "",
                "Year": "e.g. 2022",
            },
        "Chlorination Evolution Reaction":
            {
                "Formula": "e.g. Pt1(3)/CNT",
                "Type": "",
                "Elements": "e.g. Pt",
                "Content": "e.g. Pt=3",
                "Structural Property": "e.g. Pd3Pt1 (111)",
                "Electrolytes": "e.g. 0.1 M HClO4+1 M NaCl",
                "pH": "e.g. 1.3",
                "Over potential (mV@j mA cm-2)": "e.g. 50@10",
                "Cl2 selectivity": "99.5",
                "Year": "e.g. 2023",
            },
        "Ozone Evolution Reaction":
            {
                "Formula": "e.g. PtNi",
                "Type": "",
                "Elements": "e.g. Pt,Ni,B",
                "Content": "e.g. Pt/Ni/B=2.5/5.58/20.1",
                "Structural Property": "e.g. B13C2‑Encapsulated PtNi Alloy Electrocatalyst (001) (110)",
                "Electrolytes": "e.g. K2SO4",
                "pH": "e.g. 7",
                "On-set Potential (V/RHE@j mA cm-2)": "e.g. 2.54@10",
                "Tafel_slope(mVdev-1)": "e.g. 184",
                "faradaic Efficiency (%@j mAcm-2)": "e.g. 15%@50",
                "other production": "e.g. H2O2",
                "Selectivity for ozone (%@V/RHE)": "e.g. 40%",
                "Selectivity for other production (%@V/RHE)": "e.g. 90%@0.6",
                "Production rate (mg/L@j mA cm-2)": "e.g. 3.8@50",
                "Stability Test (hour@U V/RHE)": "e.g. 120@2.8",
                "Durability Loss %(performace metrics)": "e.g. 1%(Current_density)",
                "Year": "e.g. 2023",

            },
        "Nitrogen Oxidation Reaction":
            {
                "Formula": "e.g. Sr0.9RuO3",
                "Type": "",
                "Elements": "e.g. Sr,Ru,O",
                "Content": "e.g. Sr/Ru/O=0.9/1/3",
                "Structural Property": "e.g. Sr0.9RuO3(121)",
                "Electrolytes": "e.g. N2-saturated 0.1 M Na2SO4",
                "pH": "e.g. 7",
                "Quantification method": "e.g. spectroscopy",
                "Main product": "",
                "faradaic Efficiency (%@U V/RHE)": "38.6@2.2",
                "Rate (μg h-1 mg-1 @U V/RHE)": "17.9@2.2",
                "Year": "e.g. 2024",
            },
        "NH3 Oxidation Reaction":
            {
                "Formula": "e.g. CuCo/CC",
                "Type": "",
                "Elements": "e.g. Cu,Co",
                "Content": "e.g. Cu/Co=1/2",
                "Structural Property": "e.g.  self-supported CuM on carbon cloth",
                "Electrolytes": "e.g. 0.5 M K2SO4",
                "pH": "e.g. 11",
                "initial NH3 concentration (mol/L)": "e.g. 0.1",
                "Main product": "",
                "faradaic Efficiency (%@U V/RHE)": "e.g. 90%@1.85",
                "Temperature (Degrees Celsius)": "e.g. 25",
                "Year": "e.g. 2023",
            },
        "Urea Oxidation Reaction":
            {
                "Formula": "e.g. Ru-Co DAS/NiO",
                "Type": "",
                "Elements": "e.g. Ru,Co,Ni,O",
                "Content": "e.g. Ru/Co=1/1",
                "Structural Property": "e.g.  Ru-Co DAS/NiO heterostructure.NiO(222),(220),(200)",
                "Electrolytes": "e.g. 0.1 M KOH",
                "pH": "e.g. 13",
                "Main product": "",
                "faradaic efficiency (%@U V/RHE)": "e.g. 85%@1.288",
                "Potential (V/RHE@j mAcm-2)": "e.g. 1.288@10",
                "Stability (h@j mA cm-2)": "330@10",
                "Year": "e.g. 2023",
            },
        "Hydrazine Oxidation Reaction":
            {
                "Formula": "e.g. Ir-SA/NC",
                "Type": "",
                "Elements": "e.g. Ir",
                "Content": "e.g. Ir=0.45",
                "Structural Property": "e.g.  Iridium Single Atoms",
                "Electrolytes": "e.g. 0.5 M H2SO4",
                "pH": "e.g. 1.3",
                "Main product": "e.g. N2",
                "Over Potential (mV@j mA cm-2)": "e.g. 390@10",
                "Tafel slope (mV dec-1)": "e.g. 78",
                "Stability (h@j mA cm-2)": "e.g. 100@100",
                "Year": "e.g. 2023",
            },
        "Urea Synthesis":
            {
                "Formula": "e.g. In(OH)3-S",
                "Type": "",
                "Elements": "e.g. In",
                "Content": "e.g. In/O/H=1/3/3",
                "Structural Property": "e.g.  In(OH)3 (100)",
                "Electrolytes": "e.g. 0.1M KNO3+0.05 KHCO3",
                "pH": "e.g. 6.8",
                "faradaic efficiency (%@U V/RHE)": "e.g. 53.4@-0.6",
                "Rate (μg h-1 mg-1 )": "e.g. 533.1",
                "Stability (h@U V/RHE)": "12@-0.6",
                "Year": "e.g. 2023",
            },
        "Nitrate Reduction":
            {
                "Formula": "e.g. PdCu/C",
                "Type": "",
                "Elements": "e.g. Pd,Cu",
                "Content": "e.g. Pd/Cu=1/9",
                "Structural Property": "e.g.  bimetallic nanoparticles on graphene",
                "Electrolytes": "e.g. 0.2 M KOH + 0.1 M KNO3",
                "pH": "e.g. 13",
                "Current density (mA cm-2@U V/RHE)": "e.g. 5@-0.35",
                "faradaic efficiency (%@U V/RHE)": "e.g. 53.4",
                "Main product": "e.g. N2",
                "Rate (mmol min−1 gcat−1@U V/RHE)": "0.66@0.26",
                "Year": "e.g. 2023",
            },
        "Electrocatalytic hydrogenation":
            {
                "Formula": "e.g. Cu/C",
                "Type": "",
                "Elements": "e.g. Cu,C",
                "Content": "e.g. Cu/C=9.6/90.4",
                "Structural Property": "e.g. Cu/C composite",
                "Electrolytes": "e.g. 0.5 M solution of Na2SO4",
                "pH": "e.g. 13",
                "Turnover frequencies (min-1 @U V/RHE)": "0.22@-1.5",
                "faradaic efficiency (%@U V/RHE)": "e.g. 3.5@-1.5",
                "Reactant": "e.g. hydroxyacetone",
                "Main product": "e.g. propylene glycol",
                "Year": "e.g. 2016",
            },
        "CH4 Oxidation Reaction":
            {
                "Formula": "e.g. Ru-NiO/V2O5",
                "Type": "",
                "Elements": "e.g. V,Ni,Rh",
                "Content": "e.g. V/Ni=10.20/3.90",
                "Structural Property": "e.g. Rh single atoms, dispersed in NiO and V2O5",
                "Electrolytes": "e.g. 0.5 M Na2CO3",
                "pH": "e.g. 12",
                "Product": "e.g. CH3OH or CH3CH2OH",
                "faradaic efficiency (%@U V/RHE)": "e.g. 81%@1.5",
                "Rate (mol h−1 g−1)": "e.g. 0.65",
                "Year": "e.g. 2023",
            },
        "Methanol Ethanol Reforming": {
            "Formula": "e.g. Pd/C",
            "Type": "",
            "Elements": "e.g. Pd,C",
            "Content": "e.g. Pd=20",
            "Structural Property": "e.g. Pd/C",
            "Electrolytes": "e.g. 4 M KOH",
            "pH": "e.g. 13.6",
            "Reactant": "e.g. CH3CH2OH or CH3OH",
            "Product": "e.g. H2 or acetate or formate",
            "faradaic efficiency (%@U V/RHE)": "e.g. 100%@1.4",
            "Rate (mmol cm-2 h-1@U V/RHE)": "0.34@1.4",
            "Year": "e.g. 2020",
        },
        "Organic Electrocatalysis":
            {
                "Formula": "e.g. Fe–N–C-700",
                "Type": "",
                "Elements": "e.g. Fe",
                "Content": "e.g. Fe=1.71",
                "Structural Property": "e.g. Fe-pyridine/pyrrole-N M-N-C catalysts",
                "Electrolytes": "e.g. 0.5 M OA + 0.5 M NaNO3",
                "pH": "e.g. 1.3",
                "Reaction Type": "e.g. Hydrodechlorination",
                "Reactant": "e.g. Oxalic acid and NO3–/NOx",
                "Production": "e.g. glycine",
                "Current density (mA cm-2@ V/RHE)": "e.g. 10@1.5",
                "faradaic efficiency (%@U V/RHE)": "e.g. 64.2%@-0.9",
                "Mass Activity (mA mg-l@V/RHE)": "e.g. 50@-0.9",
                "ECSA(HUPD (m2 g-1)": "e.g. 0.25",
                "Specific Acticity(mA cm-2@V/RHE)": "4.34@0.9",
                "Durability Loss %(Performace metrics)": "13.7%(MA)",
                "Stability (h@U V/RHE)": "12@-0.9",
                "Year": "e.g. 2020",
            },
        "CO Oxidation Reaction":
            {
                "Formula": "e.g. Ru-NiO/V2O5",
                "Type": "",
                "Elements": "e.g. V,Ni,Rh",
                "Content": "e.g. V/Ni=10.20/3.90",
                "Structural Property": "e.g. Rh single atoms, dispersed in NiO and V2O5",
                "Electrolytes": "e.g. 0.5 M Na2CO3",
                "pH": "e.g. 12",
                "Product": "e.g. CH3OH or CH3CH2OH",
                "Onset potential (V/RHE@ mAcm-2)": "e.g. 0.03@10",
                "Current density (mAcm-2@V/RHE)": "e.g. 50@0.4",
                "faradaic efficiency (%@U V/RHE)": "e.g. 81%@1.5",
                "Rate (mol h−1 g−1)": "e.g. 0.65",
                "Year": "e.g. 2023",
            },
        "5-hydroxymethylfurfural (HMF) Oxidation":
            {
                "Formula": "e.g. Ru@NixCo1−x(OH)2",
                "Type": "",
                "Elements": "e.g. Ru,Ni,Co",
                "Content": "e.g. Ru=0.06 mmol or Ru/Ni/Co=0.1/1/1",
                "Structural Property": "e.g. Ru nanoparticles on (011) plane of Ni(OH)2 and Co(OH)2",
                "Electrolytes": "e.g. 1 M KOH",
                "pH": "e.g. 13",
                "Product": "e.g. 2,5-furandicarboxylic acid (FDCA)",
                "faradaic efficiency (%@U V/RHE)": "e.g. 99%@1.45",
                "On-set potential (V/RHE@j mA cm-2)": "e.g. 1.45@10",
                "Stability (cycles@U V/RHE)": "e.g. 6@1.40",
                "Year": "e.g. 2023",
            },
        "Acetylene semihydrogenation":
            {
                "Formula": "e.g. Cu NDs",
                "Type": "",
                "Elements": "e.g. Cu,C",
                "Content": "e.g. Cu/C=29/71 at.%",
                "Structural Property": "e.g. an average size of 4.4 ± 0.6 nm Cu nanodots, (111) plane",
                "Electrolytes": "e.g. 1 M KOH",
                "pH": "e.g. 13",
                "Main product": "e.g. C2H4",
                "faradaic efficiency (%@U V/RHE)": "e.g. 95.9%@-0.69",
                "Current density (mA cm-2@U V/RHE)": "e.g. -452@-0.79",
                "Year": "e.g. 2023",
            },
        "Ammoximation Reaction":
            {
                "Formula": "e.g. Cu NDs",
                "Type": "",
                "Elements": "e.g. Cu,C",
                "Content": "e.g. Cu/C=29/71 at.%",
                "Structural Property": "e.g. an average size of 4.4 ± 0.6 nm Cu nanodots, (111) plane",
                "Electrolytes": "e.g. 1 M KOH",
                "pH": "e.g. 13",
                "Main product": "e.g. C2H4",
                "faradaic efficiency (%@U V/RHE)": "e.g. 95.9%@-0.69",
                "Current density (mA cm-2@U V/RHE)": "e.g. -452@-0.79",
                "Yield rate (mmol/(h^-1 g^-1) @ U/RHE)": "e.g. 20@0.1",
                "Stability test (h@ mA)": "e.g. 50@60",
                "Year": "e.g. 2023",
            },
        "Oxygen-containing radical synthesis": {
            "Formula": "e.g. TiO2/C",
            "Type": "",
            "Elements": "e.g. Ti,C",
            "Content": "e.g. Ti/C=60/40 at.%",
            "Structural Property": "e.g. TiO2/C modified glassy carbon",
            "Electrolytes": "e.g. 0.05 M Na2SO4",
            "pH": "e.g. 6.8",
            "Product radicals": "e.g. HO*",
            "faradaic efficiency (%@U V/RHE)": "e.g. 95.9%@-0.69",
            "Production rate (ug min-1 cm-2)": "e.g. 2.69",
            "Current density (mA cm-2@ V/RHE)": "e.g. -0.3@-0.7",
            "Stability (h@mA cm-2)": "e.g. 100@0.1",
            "Year": "e.g. 2020",
        }
        ,
        # thermo catalysis
        "Propane dehydrogenation":
            {
                "Formula": "e.g. ZnOx",
                "Type": "",
                "Elements": "e.g. Zn,O",
                "Content": "e.g.Zn/O:1/1",
                "Structural Property": "e.g. a physical mixture of ZnO (8wt%) and non-acidic SiO2 material",
                "Temperature(Celsius)": "e.g. 550",
                "Pressure (bar)": "1.2",
                "Reactant": "e.g. C3H8:H2:N2=4:2:4",
                "rate (mmol-1 g-1 min-1)": "e.g. 1.7",
                "Conversion degree (%)": "e.g. 10%",
                "TOF (h-1)": "e.g. 239",
                "Selectivity (%)": "e.g. 91.9",
                "Stability (h)": "e.g. 400",
                "Year": "e.g. 2021"
            },
        "Carbon monoxide oxidation":
            {
                "Formula": "e.g. ZnOx",
                "Type": "",
                "Elements": "e.g. Zn,O",
                "Content": "e.g. Zn/O:1/1",
                "Structural Property": "e.g. a physical mixture of ZnO (8wt%) and non-acidic SiO2 material",
                "Temperature(Celsius@conversion 10%)": "e.g. 550",
                "Temperature(Celsius)": "e.g. 550",
                "Pressure (bar)": "1.2",
                "Reactant": "e.g. CO:O2=2:1",
                "rate (mol g-1 h-1)": "e.g. 1.7",
                "Conversion degree (%)": "e.g. 10%",
                "TOF (s-1)": "e.g. 2.5",
                "Apparent activation energy Ea (kJ/mol)": "e.g. 2.4",
                "Selectivity (%)": "e.g. 91.9",
                "CO Reaction order": "e.g. 2",
                "O2 Reaction order": "e.g. 1",
                "Stability (h)": "e.g. 100",
                "Year": "e.g. 2021"
            },
        "Carbon dioxide hydrogenation":
            {
                "Formula": "e.g. ZnOx",
                "Type": "",
                "Elements": "e.g. Zn,O",
                "Content": "e.g. Zn/O=1/1",
                "Structural Property": "e.g. a physical mixture of ZnO (8wt%) and non-acidic SiO2 material",
                "Temperature(Celsius@conversion 10%)": "e.g. 550",
                "Temperature(Celsius)": "e.g. 550",
                "Pressure (bar)": "e.g. 1.2",
                "Reactant": "e.g. CO:O2=2:1",
                "pressure of CO2 (bar)": "e.g. 0.7",
                "pressure of H2 (bar)": "e.g. 0.5",
                "pressure of CO (bar)": "e.g. 0.2",
                "rate (g Kgcat-1 h-1)": "e.g. 1.7",
                "CH3OH selectivity (%)": "e.g. 91.9",
                "CO selectivity (%)": "e.g. 91.9",
                "CO2 Conversion (%)": "e.g. 91.9",
                "CO Conversion (%)": "e.g. 91.9",
                "particle size(nm)": "e.g. 20",
                "TOF (s-1)": "e.g. 10",
                "Apparent activation energy Ea (kJ/mol)": "e.g. 100",
                "Stability (h)": "e.g. 100",
                "CO2 Reaction order": "e.g. 2",
                "H2 Reaction order": "e.g. 1",
                "CO Reaction order": "e.g. 2",
                "CH3OH Reaction order": "e.g. 2",
                "Year": "e.g. 2024",
                "last corresponding author": "e.g. Ang Cao, Hao Li"
            },
        "Thermal Ammonia Synthesis":
            {
                "Formula": "e.g. ZnOx",
                "Elements": "e.g. Zn,O",
                "Content": "e.g. Zn/O=1/1",
                "Structural Property": "e.g. a physical mixture of ZnO (8wt%) and non-acidic SiO2 material",
                "Temperature(Celsius)": "e.g. 550",
                "total Pressure (bar)": "e.g. 1.2",
                "Reactant ratio": "e.g. CO:O2=2:1",
                "pressure of N2 (bar)": "e.g. 0.7",
                "pressure of H2 (bar)": "e.g. 0.7",
                "gas space velocity (mL/h)": "e.g. 20",
                "NH3 formation rate (g Kgcat-1 h-1)": "e.g. 27",
                "N2 Conversion (%)": "e.g. 91.9",
                "particle size(nm)": "e.g. 20",
                "TOF (s-1)": "e.g. 10",
                "Apparent activation energy Ea (kJ/mol)": "e.g. 100",
                "Stability (h)": "e.g. 100",
                "N2 Reaction order": "e.g. 2",
                "H2 Reaction order": "e.g. 1",
                "NH3 Reaction order": "e.g. 2",
                "Year": "e.g. 2024",
                "last corresponding author": "e.g. Ang Cao, Hao Li"
            },
        "Carbon dioxide + methonal = DMC(dimethyl carbonate)":
            {
                "Formula": "e.g., P-CeO2-NR",
                "Elements": "e.g., Ce,O",
                "Content": "e.g., AC-Fe=0.1838:1",
                "Structural Property": "e.g., cubofluorite",
                "Dehydrating agent": "e.g., 2-CP",
                "methanol/dehydrating agent (mol/mol)": "e.g., 21.3",
                "Temperature(Celsius)": "e.g., 139.85",
                "Pressure (MPa)": "e.g., 2",
                "rate (mmol Kgcat-1 h-1)": "e.g, 11.9",
                "DMC selectivity (%)": "e.g., 100",
                "CH3OH Conversion (%)": "e.g., 5.5",
                "DMC Yield (%)": "e.g., 23.5",
                "particle size(nm)": "e.g., 11",
                "Stability (h)": "e.g., 24",
                "Year": "e.g., 2024",
                "last corresponding author": "e.g., Hao Li"
            },
        "Fischer-Tropsch Synthesis":
            {
                "Formula": "e.g., Co/Al2O3",
                "Pore size (nm)": "e.g., 8",
                "Pore volume (m3g-1)": "e.g., 0.34",
                "Elements": "e.g., Co, Al, O",
                "Content": "e.g., Co=20 wt%",
                "BET surface area (m2g-1)": "e.g., 139",
                "Dispersion (%)": "e.g., 4.4",
                "Co loading (%)": "e.g., 20",
                "Particle size (nm)": "e.g., 9.3",
                "Degree of reduction (%)": "e.g., 28.4",
                "Reduction temperature (K)": "e.g., 673",
                "Catalyst support": "e.g., Al2O3",
                "Catalyst promoter": "e.g., Ba",
                "H2/CO": "e.g., 2",
                "Temperature(K)": "e.g., 498",
                "Pressure(Mpa)": "e.g., 2",
                "Time on stream(h)": "e.g., 48",
                "GHSV(h-1)": "e.g., 1000",
                "GHSV(cm3g-1h-1)": "e.g., 1000",
                "CO conversion rate": "e.g., 58.4",
                "TOF(s-1)": "e.g., 0.0189",
                "C1 hydrocarbons selectivity (%)": "e.g., 16.1",
                "C5+hydrocarbons selectivity": "e.g., 67.6",
                "Year": "e.g., 2024"
            },
        "Methanol Oxidative Carbonylation to DMC":
            {
                "Formula": "e.g. Cu-N-C",
                "Elements": "e.g., Cu, C, N, O",
                "Content": "e.g. Cu = 7.0 or Cu/C = 7/93",
                "Structural Property": "e.g. Single Cu atoms on nitrogen doped graphene",
                "Device": "e.g. Slurry bed or Fixed bed",
                "Temperature (Celsius)": "e.g. 120",
                "Time (h)": "e.g., 1.5",
                "total Pressure (bar)": "e.g., 30",
                "Mass of catalyst (mg)": "e.g., 100",
                "Volume of CH3OH (mL)": "e.g., 7.5",
                "pressure of CO (bar)": "e.g., 20",
                "pressure of O2 (bar)": "e.g., 10",
                "Volume flow rate of CH3OH (mL/min)": "e.g., 0.02",
                "Volume flow rate of CO (mL/min)": "e.g., 26",
                "Volume flow rate of O2 (mL/min)": "e.g., 2",
                "CH3OH Conversion (%)": "e.g., 3.1",
                "DMC selectivity": "e.g., 97.6",
                "CO Conversion (%)": "e.g., 14",
                "O2 Conversion (%)": "e.g., 20",
                "Space-time yield of DMC (g·gcat-1·h-1)": "e.g., 1.67",
                "Space-time yield of DMC (g·gmetal-1·h-1)": "e.g., 23.8",
                "particle size(nm)": "e.g., 7.5",
                "TOF (h-1)": "e.g., 16.8",
                "Average deactivation every 10 hours (%)": "e.g., 4.3",
                "Average deactivation per cycle (%)": "e.g., 2.5",
                "last corresponding author": "e.g., Hao Li",
                "Year": "2018"
            },

        # photocatalysis
        "Solar Water Splitting": {
            "Formula": "e.g., TiO2@Pt,RuO2",
            "Elements": "e.g., Ti,Pt,Ru,O",
            "Content": "e.g. Ti/Pt = 1:1",
            "Structural Property": "e.g., titanium dioxide (TiO₂) (001)face is combined with platinum (Pt) and ruthenium dioxide (RuO₂).",
            "pH": "e.g., 1.5",
            "bandgap (eV)": "e.g., 3.2",
            "absorption range (nm)": "e.g., 385",
            "Reactant": "e.g., water",
            "Light source": "e.g., 450-W Xe lamp",
            "H2 evolution rate (umol/h)": "e.g., 2.8",
            "O2 evolution rate (umol/h)": "e.g., 1.4",
            "reaction duration (h)": "e.g., 25",
            "quantum efficienty at light wavelength (%@nm)": "e.g., 30@10",
            "Apparent quantum yield at light wavelength (%@nm)": "e.g., 20@10",
            "Year": "e.g., 2024",
            "last corresponding author": "e.g., Michael Gratze",
            "Corresponding author institute": "e.g., Tohoku university"
        }

    }

    def add_key_value_after(dictionary, key, new_key, new_value):
        keys = list(dictionary.keys())
        if key in keys:
            index = keys.index(key)
            new_dict = {}
            for i, k in enumerate(keys):
                new_dict[k] = dictionary[k]
                if i == index:
                    new_dict[new_key] = new_value
            return new_dict
        else:
            raise KeyError(f"Key '{key}' not found in the dictionary.")

    if mag_flag:
        for k, v in _column_title_dict.items():
            if "Electrolytes" in v.keys():
                _column_title_dict[k] = add_key_value_after(v, "Electrolytes", "Magnetic field (mT)", "e.g. 100")
    if sea_flag:
        for k, v in _column_title_dict.items():
            if "Electrolytes" in v.keys() and "Tafel_slope (mVdev-1)" in v.keys():
                _column_title_dict[k] = add_key_value_after(_column_title_dict[k], "Electrolytes",
                                                            "Seawater simulated electrolyte", "e.g. 1M NaCl")
                _column_title_dict[k] = add_key_value_after(_column_title_dict[k], "Tafel_slope (mVdev-1)",
                                                            "Overall seawater spilting Over_Potential (mV@ j mA cm-2)",
                                                            "e.g. 234@10")
                _column_title_dict[k] = add_key_value_after(_column_title_dict[k],
                                                            "Overall seawater spilting Over_Potential (mV@ j mA cm-2)",
                                                            "Overall seawater spilting Current_density (mA/cm2@ U V/RHE)",
                                                            "e.g. 40@1.79")
                _column_title_dict[k] = add_key_value_after(_column_title_dict[k],
                                                            "Overall seawater spilting Current_density (mA/cm2@ U V/RHE)",
                                                            "Overall seawater spilting Stability Test (h@ mAcm-2)",
                                                            "e.g. 200@500")
    return _column_title_dict[reaction_name]


def special_metric(reaction_name, key):
    special_dict = {
        "Carbon Dioxide Reduction": {"Main product": True, "Other product (FE > 10%)": True},
        "Ammonia Synthesis": {"Isotope labeling": True, "Reactant": True},
        "Epoxide Production": {"Main product": True},
        "Nitrogen Oxidation Reaction": {"Main product": True},
        "NH3 Oxidation Reaction": {"Main product": True},
        "Urea Oxidation Reaction": {"Main product": True}
    }

    # Use a single line to get the value from the nested dictionary
    return special_dict.get(reaction_name, {}).get(key, False)


def subcategory_dict(main_category):
    sub_dict = {
        "Platinum/Precious_Group_Metal": ["Pt-based", "Ru-based", "Rh-based", "Pd-based",
                                          "Os-based", "Ir-based", "Au-based", "Ag-based"],
        "Metal/Alloy": ["Transition Metal/Alloy", "Main group Metal/Alloy"],
        "Metal_C/N/O/HO/F/S/P-ides": ["Carbide", "Nitride", "Oxide", "Hydroxide",
                                      "Fluoride", "Sulfide", "Phosphide"],
        "Metal-Nitrogen-Carbon(CNT/graphene)": ["Single atom catalysts (pyridine/pyrrole-N)",
                                                "Single atom catalysts (COF/MOF)",
                                                "Single-atom catalysts with other elemental coordination",
                                                "Dual atom catalysts", "Metal/Nitrogen-doped CNT",
                                                "Metal/Nitrogen-doped graphene"],
        "2D_materials": ["graphene/graphene oxide", "MXene", "Transition metal sulphides(TMDs)", "Borophene/BN"],
        "Perovskite": [],
        "Others": []
    }
    return sub_dict[main_category]


def write_excel(df, filename, sheet_name, append=True):
    # Ensure the path to the file is correct if you're using relative paths
    cwd = os.getcwd()
    full_path = os.path.join(cwd, filename)

    # Use try-except to handle the case where the file does not exist
    try:
        # Use mode='a' to append if the file already exists
        with pd.ExcelWriter(full_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # Load the existing workbook if it exists
            if not writer.book:
                writer.book = load_workbook(full_path)
                writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

            if sheet_name in writer.sheets and append:
                # If the sheet exists, read the old data and append the new data
                # Load existing data into a DataFrame
                df_old = pd.read_excel(full_path, sheet_name=sheet_name)
                # Concatenate the old data with the new data
                df_final = pd.concat([df_old, df], ignore_index=True)
            else:
                # If the sheet does not exist, use the new DataFrame directly
                df_final = df

            # Write the final DataFrame to the specified sheet
            df_final.to_excel(writer, sheet_name=sheet_name, index=False)

    except FileNotFoundError:
        # If the file does not exist, create a new one
        df.to_excel(full_path, sheet_name=sheet_name, index=False)
    return True


def doi_contributed_by_user(doi, reaction_name, drive_client, excel_dir="User_contribution"):
    cwd = os.getcwd()
    excel_list = os.listdir(os.path.join(cwd, excel_dir))
    if len(excel_list) > 0:
        for excel_file in excel_list:
            df = read_excel_data(os.path.join(cwd, excel_dir, excel_file), sheet_name=reaction_name)
            if not df.empty:
                if doi in df["DOI"].values:
                    return True
        return False
    else:
        return False


def get_doi(doi_input, reaction_name, _dic):
    for reaction in _dic.keys():
        if reaction == sheet_dict[reaction_name]:
            mask_is_null = _dic[reaction]["modify time"].isnull()
            _dic = _dic[reaction][mask_is_null]
            _dic["DOI"] = _dic["DOI"].str.strip()
            if doi_input.strip() in _dic["DOI"].values:
                return None
    return doi_input


def get_doi_db(_dic):
    _doi_ls = []
    for sheet_name in _dic.keys():
        mask_is_null = _dic[sheet_name]["modify time"].isnull()
        _dic[sheet_name] = _dic[sheet_name][mask_is_null]
        _doi_ls += (list(set(_dic[sheet_name]["DOI"].str.strip())))
    return _doi_ls


def extract_and_modify_data(reaction_name, name, username, institute, thermo_flag, photo_flag, admin=None,
                            admin_key="DOI", admin_value=""):
    total_excel = get_total_excel(thermo_flag, photo_flag)
    df_to_be_modified = total_excel[reaction_name]
    st.info("Please do not add new catalys here. This form is only used for modification!")
    # st.write(df_to_be_modified)
    if username in admin:
        mask = (df_to_be_modified[admin_key] == admin_value) & (df_to_be_modified["modify time"].isnull())
        # st.write(admin_key)
        # st.write(admin_value)
        # st.write(mask)
    else:
        mask = ((df_to_be_modified['Name'] == name) | (df_to_be_modified['Name'] == username)) & (
                    df_to_be_modified['Institute'] == institute) & (df_to_be_modified["modify time"].isnull())
    extracted_data = df_to_be_modified[mask]

    if len(extracted_data) > 0:
        st.session_state.modified_data = extracted_data
        st.session_state.modified_data = st.data_editor(st.session_state.modified_data, num_rows="dynamic",
                                                        column_config={"modify time": None},
                                                        disabled=("Name", "Institute"), hide_index=False)
        if st.button("Save changes and continue to upload"):
            with st.status("Updating data...may take 5-10 seconds..", expanded=True) as status:
                if check_and_update_time():
                    st.write("Generating new data files...")
                    df_to_be_modified.loc[mask, "modify time"] = pd.Timestamp('now')
                    df_to_be_modified = pd.concat([df_to_be_modified, st.session_state.modified_data],
                                                  ignore_index=True)
                    total_excel[reaction_name] = df_to_be_modified
                    st.write("Updating Cloud Data...")
                    write_total_pickle_upload(total_excel, thermo_flag, photo_flag)
                    status.update(label="Update complete!", state="complete", expanded=False)
                    st.session_state["modify_data"] = False
                    st.rerun()
                else:
                    st.error('System busy, upload has failed, please try again after 10 seconds!')
        if st.button("Give up changes and quit"):
            with st.status("Return to upload entry...", expanded=False) as status:
                st.session_state["modify_data"] = False
                status.update(label="Complete!", state="complete", expanded=False)
                st.rerun()
    else:
        st.warning("No matching data found, returning to upload...")
        time.sleep(1)
        st.session_state["modify_data"] = False
        if st.button("Return"):
            st.session_state["modify_data"] = False
        return False


def write_total_pickle_upload(df, thermo_flag, photo_flag):
    if not thermo_flag and not photo_flag:
        pickle_path = os.path.join(".", "total_excel", "total.pkl")
        with open(pickle_path, 'wb') as file:
            pickle.dump(df, file)
        upload_or_replace_file("total.pkl", pickle_path, "application/octet-stream", dir_dict["total_excel"])
    elif thermo_flag:
        pickle_path = os.path.join(".", "total_pickle_thermocatalysis", "thermo_excel.pkl")
        print(pickle_path)
        with open(pickle_path, 'wb') as file:
            pickle.dump(df, file)
        upload_or_replace_file("thermo_excel.pkl", pickle_path, "application/octet-stream",
                               dir_dict["total_pickle_thermocatalysis"])
    elif photo_flag:
        pickle_path = os.path.join(".", "total_pickle_photocatalysis", "photo_excel.pkl")
        with open(pickle_path, 'wb') as file:
            pickle.dump(df, file)
        upload_or_replace_file("photo_excel.pkl", pickle_path, "application/octet-stream",
                               dir_dict["total_pickle_photocatalysis"])


def check_and_update_time():
    file_path = "last_saved_time.txt"
    current_time = time.time()

    try:
        with open(file_path, "r") as file:
            last_saved_time = float(file.read())
    except (FileNotFoundError, ValueError):
        last_saved_time = 0

    if current_time - last_saved_time > 15:
        with open(file_path, "w") as file:
            file.write(str(current_time))
        return True
    else:
        return False


def count_uploading_data_within_time_range(t_df, start_time, end_time):
    start_time = pd.to_datetime(start_time)
    end_time = pd.to_datetime(end_time)
    count = 0
    institute_dict = {}
    name_dict = {}
    for sheet_name, df in t_df.items():

        # 检查当前sheet中是否存在名为“Uploading data”的列
        df = df[df['modify time'].isna()]
        st.write(sheet_name)
        if 'Uploading data' in df.columns:
            # 将“Uploading data”列转换为时间对象
            df['Uploading data'] = pd.to_datetime(df['Uploading data'], errors='coerce')
            # 过滤符合时间范围的条目
            filtered_df = df[(df['Uploading data'] >= start_time) & (df['Uploading data'] <= end_time)]
            count += filtered_df.shape[0]
            st.write(filtered_df)
            st.write(count)

        for ins in filtered_df['Institute']:
            if not ins in institute_dict.keys():
                institute_dict[ins] = 1
            else:
                institute_dict[ins] += 1

        for name in filtered_df['Name']:
            if not name in name_dict.keys():
                name_dict[name] = 1
            else:
                name_dict[name] += 1
    st.write(institute_dict)
    st.write(name_dict)
    st.write(count)
    return count


def get_expected_reactions(thermo, photo) -> dict:
    if thermo:
        return column_title_dict["thermo"]
    elif photo:
        return column_title_dict["photo"]
    else:
        return column_title_dict["electro"]


if __name__ == "__main__":
    if not "modify_data" in st.session_state:
        st.session_state["modify_data"] = False

    reaction_name, mag_flag, thermo_flag, sea_flag, photo_flag = streamlit_frame()
    total_dic = get_total_excel(thermo_flag, photo_flag)
    # st.write(total_dic["ORR"]["DOI"])
    doi_database = get_doi_db(total_dic) # 当前flag下，存在的所有doi
    expected_reactions = get_expected_reactions(thermo_flag, photo_flag)
    current_df = total_dic[sheet_dict[reaction_name]]

    st.markdown("---")
    cwd = os.getcwd()

    new_df = {}
    # col1, col2 = st.columns(2)

    yaml_file_path = os.path.join(cwd, "user.yaml")
    if not os.path.exists(yaml_file_path):
        raise "Can not found file user.yaml"
        file_from_gdrive(dir_dict["user"], "user.yaml", init_drive_client(), yaml_file_path)
    login_result = login(yaml_file=os.path.join(cwd, "user.yaml"))
    # login_result = login("./user.yaml")

    if login_result is not None:
        authentication_status, name, username, institute = login_result
    else:
        authentication_status, name, username, institute = (None, "", "", "")

    if authentication_status is None:
        def welcome():
            st.markdown("<div style='text-align: center;'>"
                        "<h1 style='font-size: 32px;'>👋 Welcome to DigCat upload system!</h1>"
                        "</div>", unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)

            with st.expander("📚 Tips", expanded=True):
                st.markdown("[<b>Digital Catalysis Platform </b>:](https://digcat.streamlit.app/)",
                            unsafe_allow_html=True)
                st.markdown("""
                    <div style='text-align: justify; font-size: 24px;'>
                    <ul style='font-size: 24px;'>
                    <li>Data Analytics 📊</li>
                    <li>Modeling 🧮</li>
                    <li>AI Learning 🧠</li>
                    <li>Performance Benchmarking 📈</li>
                    <li>And many other functions! ➕</li>
                    </ul>
                    </div>
                    <div style='text-align: justify; font-size: 20px;'>
                    To get started, please <span style='color: #ff4b4b;'>
                    <b>register</b>
                    </span> or <span style='color: #ff4b4b;'>
                    <b>login</b>
                    </span> using the form on the left sidebar.
                    </div>
                """, unsafe_allow_html=True)


        welcome()

    admin = ["tohokudizhang"]
    if authentication_status is not None:
        st.sidebar.markdown("#### :point_down: Need to modify uploaded data? ")
        if username in admin:
            admin_key = st.sidebar.text_input('Please input key', '', label_visibility="collapsed")
            admin_value = st.sidebar.text_input('Please input value', '', label_visibility="collapsed")

            if st.sidebar.button("check total excel"):
                total_df = get_total_excel(thermo_flag, photo_flag)
                st.write(total_df[sheet_dict[reaction_name]])
            st.title("Count Uploading Data Within Time Range")
            # Date input using st.date_input
            today = mydt.datetime.now()
            jan_1 = mydt.date(today.year, 1, 1)
            dec_31 = mydt.date(today.year, 12, 31)
            d = st.date_input(
                "Select the date range",
                (mydt.date(today.year, today.month, 1), mydt.date(today.year, today.month, today.day)),
                jan_1,
                dec_31,
                format="MM.DD.YYYY",
            )
            if st.button("Count Data"):
                total_df = get_total_excel(thermo_flag, photo_flag)
                count_uploading_data_within_time_range(total_df, str(d[0]) + " 00:00:00", str(d[1]) + " 00:00:00")
            if st.sidebar.button("clean electro pickle"):
                if os.path.exists(os.path.join(".", "total_excel", "total.pkl")):
                    os.remove(os.path.join(".", "total_excel", "total.pkl"))
            if st.sidebar.button("clean thermo pickle"):
                if os.path.exists(os.path.join(".", "total_pickle_thermocatalysis", "thermo_excel.pkl")):
                    os.remove(os.path.join(".", "total_pickle_thermocatalysis", "thermo_excel.pkl"))
        else:
            admin_key = ""
            admin_value = ""
        if st.sidebar.button("Modify uploaded data"):
            st.session_state["modify_data"] = True

        if st.session_state["modify_data"]:
            extract_and_modify_data(sheet_dict[reaction_name], name, username, institute, thermo_flag, photo_flag,
                                    admin=admin, admin_key=admin_key, admin_value=admin_value)
        else:
            # df = read_excel_data(excel_file_name, sheet_name=reaction_name)
            with st.container():
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("### Upload by online interface")
                    st.markdown("##### Please enter doi：")
                    st.write('e.g. 10.1126/science.aaw7493')
                    doi_input = st.text_input(' ', '10.xxx...', label_visibility="collapsed")
                    if doi_input.startswith("https://doi.org/"):
                        doi_input = doi_input.lstrip("https://doi.org/ ")
                    elif not bool(re.match(r'^\d', doi_input)):
                        st.markdown(":red[**doi not valid**]")
                    # st.markdown('**Note**: Please click "Check" to update the status after each doi change.')
                    if 'doi_clicked' not in st.session_state:
                        st.session_state.doi_clicked = False

                    if "doi" not in st.session_state:
                        st.session_state.doi = False

                    if st.button('Check'):
                        st.session_state.doi_clicked = True
                        st.session_state['unpublished button'] = False

                    if "checked_reaction" not in st.session_state:
                        st.session_state["checked_reaction"] = None

                    if st.session_state.doi_clicked:
                        st.session_state.doi = get_doi(doi_input, reaction_name, total_dic)
                        st.session_state["checked_reaction"] = reaction_name
                        st.session_state.doi_clicked = False

                    if 'unpublished button' not in st.session_state:
                        st.session_state['unpublished button'] = False

                    # st.markdown("##### Please check for duplicates：")
                    if doi_input == "":
                        st.markdown('Please fill in the doi')
                    elif (st.session_state.doi == False and doi_input != "") or (
                            st.session_state.doi != doi_input and st.session_state.doi is not None):
                        st.markdown('Please :red[Click] to check if the paper is uploaded')
                    elif st.session_state.doi is None and doi_input != "":
                        st.markdown('This paper :red[**is in**] the database')
                        st.session_state.doi = False
                    elif (st.session_state.doi == doi_input and st.session_state.doi != None and st.session_state[
                        "checked_reaction"] == reaction_name):
                        st.markdown('This paper :red[**is not**] in the database')
                        with col2:
                            st.markdown("##### Please upload the pdf：")
                            uploaded_file = st.file_uploader("Choose a PDF file", type="pdf")
                            si_flag = st.checkbox("Please tick it if uploading the Supplementary Information (SI)")
                            si_button = st.button("Upload")
                            if si_button and uploaded_file is not None:
                                drive_client = init_drive_client()
                                if not si_flag:
                                    uploaded_file.name = f"{st.session_state.doi}.pdf"
                                else:
                                    uploaded_file.name = f"{st.session_state.doi}_si.pdf"
                                if thermo_flag:
                                    utils.save_files(uploaded_file, "./pdfs_thermocatalysis")
                                    st_upload_file_to_drive(uploaded_file, dir_dict["pdfs_thermocatalysis"], drive_client)
                                elif photo_flag:
                                    utils.save_files(uploaded_file, "./pdfs_photocatalysis")
                                    st_upload_file_to_drive(uploaded_file, dir_dict["pdfs_photocatalysis"], drive_client)
                                else:
                                    utils.save_files(uploaded_file, "./pdfs")
                                    st_upload_file_to_drive(uploaded_file, dir_dict["pdfs"], drive_client)
                                st.success('File has been uploaded!')
                            elif si_button and uploaded_file is None:
                                st.error("Please add a file to upload.")

                    # st.write('The current doi is:', st.session_state.doi)
                    if st.button("Need to upload unpublished works? Please click here"):
                        st.session_state['unpublished button'] = True
                        st.session_state.doi = "Unpublished"

                # data submission
                if (st.session_state.doi == doi_input and st.session_state.doi != None and st.session_state["checked_reaction"] == reaction_name) or st.session_state['unpublished button']:
                    if st.session_state['unpublished button']:
                        st.markdown("**Uploading unpublished articles now**")
                    new_df["DOI"] = st.session_state.doi
                    st.markdown("---")
                    st.markdown("##### Please select Type：")
                    type_col1, type_col2 = st.columns(2)
                    with type_col1:
                        main_category = st.selectbox("Main category", (
                            "Platinum/Precious_Group_Metal",
                            "Metal/Alloy",
                            "Metal_C/N/O/HO/F/S/P-ides",
                            "Metal-Nitrogen-Carbon(CNT/graphene)",
                            "2D_materials",
                            "Perovskite",
                            "Others"
                        ), label_visibility="visible")
                        new_df["Type"] = main_category
                        # st.write(new_df["Type"])
                    with type_col2:
                        if len(subcategory_dict(main_category)) > 0:
                            subcategory = st.selectbox("Subcategory", subcategory_dict(main_category), label_visibility = "visible")
                            new_df["subType"] = subcategory

                    with st.form(key='my_form'):
                        c_index = 0
                        for key, value in reaction_column_title_dict(reaction_name, mag_flag, sea_flag).items():
                            if key != "Type" and key != "Content" and not special_metric(reaction_name, key):
                                c_index += 1
                                st.markdown(f"##### {c_index}. Please enter {key}：")
                                if "(" in key and ")" in key:
                                    result_col, costum_unit = st.columns(2)
                                    with result_col:
                                        st.write(value)
                                        results = st.text_input(' ', "", key=key, label_visibility="collapsed")
                                    with costum_unit:
                                        costum_unit_flag = st.checkbox('Custom unit', key=key + "check_unit")
                                        costum_unit_result = st.text_input(' ', "", key=key + "unit",
                                                                           label_visibility="collapsed")
                                        if costum_unit_result == "" and costum_unit_flag:
                                            st.error("Please fill in the costum unit.")
                                        elif costum_unit_result != "" and costum_unit_flag:
                                            results = results + f"({costum_unit_result})"
                                else:
                                    st.write(value)
                                    results = st.text_input(' ', "", key=key, label_visibility="collapsed")
                                if "@" in key and "@" not in results and results != "" and results != "NA":
                                    st.error("formatting error")
                                    results = "formatting error"
                            elif key == "Content":
                                c_index += 1
                                st.markdown(f"##### {c_index}. Please enter {key}：")
                                ratio_col1, ratio_col2 = st.columns(2)
                                with ratio_col1:
                                    st.write(value)
                                    results = st.text_input(' ', "", key=key, label_visibility="collapsed")
                                    st.write("If there are multiple units of different component ratios:")
                                    results_more = st.text_input(' ', "", key=key + "content_more",
                                                                 label_visibility="collapsed")
                                with ratio_col2:
                                    "Slelect mass ratio (wt.%) or atomic ratio (at.%)"
                                    ratio_option = st.selectbox(' ', ('at.%', 'wt.%', "mmol", "mg"),
                                                                label_visibility="collapsed")
                                    st.write("Multiple units of component ratios:")
                                    ratio_option_more = st.selectbox(' ', ('at.%', 'wt.%', "mmol", "mg"),
                                                                     key="content_more", label_visibility="collapsed")

                                results = results + " " + ratio_option + ";" + results_more + " " + ratio_option_more
                            # special treatment
                            elif special_metric(reaction_name, key):
                                c_index += 1
                                st.markdown(f"##### {c_index}. Please select {key}：")
                                if reaction_name == "Carbon Dioxide Reduction":
                                    if key == "Main product":
                                        CO2RR_main_product = st.selectbox(' ', (
                                            'Acetate', 'Carbon monoxide(CO)', "Methanal(CH2O)",
                                            'Formate', 'Ethylene(C2H4)', 'Ethanol(CH3CH2OH)', 'Ethanal',
                                            'Methanol', "Methane(CH4)", "Hydrogen(H2)", "Ethane (C2H6)", "C3 products",
                                        ), label_visibility="collapsed")
                                        results = CO2RR_main_product
                                    elif key == "Other product (FE > 10%)":
                                        CO2RR_other_product = st.multiselect(' ', (
                                            'Acetate', 'Carbon monoxide(CO)', "Methanal(CH2O)",
                                            'Formate', 'Ethylene(C2H4)', 'Ethanol(CH3CH2OH)', 'Ethanal',
                                            'Methanol', "Methane(CH4)", "Hydrogen(H2)", "Ethane (C2H6)", "C3 products",
                                            "NA",
                                        ), label_visibility="collapsed")
                                        results = ";".join(CO2RR_other_product)
                                elif reaction_name == "Carbon Dioxide Reduction":
                                    if key == "Main product":
                                        CORR_main_product = st.selectbox(' ', (
                                            'Acetate', 'Carbon monoxide(CO)', "Methanal(CH2O)",
                                            'Formate', 'Ethylene(C2H4)', 'Ethanol(CH3CH2OH)', 'Ethanal',
                                            'Methanol', "Methane(CH4)", "Hydrogen(H2)", "Ethane (C2H6)", "C3 products",
                                        ), label_visibility="collapsed")
                                        results = CORR_main_product
                                    elif key == "Other product (FE > 10%)":
                                        CORR_other_product = st.multiselect(' ', (
                                            'Acetate', 'Carbon monoxide(CO)', "Methanal(CH2O)",
                                            'Formate', 'Ethylene(C2H4)', 'Ethanol(CH3CH2OH)', 'Ethanal',
                                            'Methanol', "Methane(CH4)", "Hydrogen(H2)", "Ethane (C2H6)", "C3 products",
                                            "NA",
                                        ), label_visibility="collapsed")
                                        results = ";".join(CORR_other_product)
                                elif reaction_name == "Ammonia Synthesis":
                                    if key == "Isotope labeling":
                                        results = st.radio(
                                            "Are isotope labelling experiments performed?",
                                            ["yes", "no"], )
                                    elif key == "Reactant":
                                        results = st.selectbox(' ', ('N2', 'NO/NO2/NO3', "NO3-/NO2-/NO-"
                                                                     ), label_visibility="collapsed")
                                elif reaction_name == "Epoxide Production":
                                    if key == "Main product":
                                        results = st.selectbox(' ', ('cyclooctene oxide',
                                                                     'propylene oxide',
                                                                     "ethylene oxide",
                                                                     'acrolein',
                                                                     'acrylic acid',
                                                                     'allyl alcohol',
                                                                     'propyiene glycol',
                                                                     'acetone',

                                                                     ), label_visibility="collapsed")
                                elif reaction_name == "Nitrogen Oxidation Reaction":
                                    if key == "Main product":
                                        results = st.selectbox(' ', ('NO3-', 'NO2-', 'NO-'
                                                                     ), label_visibility="collapsed")
                                        st.write()
                                elif reaction_name == "NH3 Oxidation Reaction":
                                    if key == "Main product":
                                        results = st.selectbox(' ', ('N2', 'N2O',
                                                                     ), label_visibility="collapsed")
                                        st.write()
                                elif reaction_name == "Urea Oxidation Reaction":
                                    if key == "Main product":
                                        urea_product = st.multiselect(' ', ('N2', 'NO-',
                                                                            'NO2-', 'NO3-', 'CO2',
                                                                            'CO'
                                                                            ), label_visibility="collapsed")
                                        results = ";".join(urea_product)
                            if key != "Type":
                                new_df[key] = results

                        # st.markdown("##### Please enter the uploader：")
                        # uploader = st.text_input('Please always use the same uploader to accumulate credits.', "e.g. xuan wang", label_visibility="visible")
                        new_df["Name"] = username
                        new_df["Institute"] = institute
                        new_df["Uploading data"] = pd.Timestamp(datetime.now())
                        _, sub_col1, sub_col2, sub_col3, _ = st.columns(5)
                        with sub_col1:
                            save_button = st.form_submit_button(label='Save & preview')
                        with sub_col3:
                            submit_button = st.form_submit_button(label='Submit')

                    if save_button:
                        error_flag = False
                        for key, value in new_df.items():
                            if value == "formatting error":
                                st.error("Please address formatting errors")
                                error_flag = True
                                break
                            if value == "":
                                st.error(f"Please fill in {key}. (Please fill in 'NA' for missing items)")
                                error_flag = True
                                break
                        if not error_flag:
                            new_df = pd.DataFrame([new_df])
                            st.table(new_df)

                    if submit_button:
                        error_flag = False
                        for key, value in new_df.items():
                            if value == "formatting error":
                                st.error("Please address formatting errors")
                                error_flag = True
                                break
                            if value == "":
                                st.error(f"Please fill in {key}. (Please fill in 'NA' for missing items)")
                                error_flag = True
                                break
                        if not error_flag:
                            new_df = pd.DataFrame([new_df])
                            total_df = get_total_excel(thermo_flag, photo_flag)
                            total_df[sheet_dict[reaction_name]] = pd.concat(
                                [total_df[sheet_dict[reaction_name]], new_df], ignore_index=True)
                            if check_and_update_time():
                                write_total_pickle_upload(total_df, thermo_flag, photo_flag)
                                st.success('Data has been submitted!')
                                st.warning(
                                    'Please make sure that all data for each paper is submitted before returning to the DOI check to verify that the article is in the database.')
                            else:
                                st.error('System busy, upload has failed, please try again after 10 seconds!')
            for i in range(6):
                st.write("\n")
            # 以下为通过Excel批量上传的代码
            with st.container():
                def upload_cur_df(df, reaction_name):
                    print("---------------------")
                    df["Name"] = name
                    df["Institute"] = institute
                    df["Uploading data"] = pd.Timestamp(datetime.now())
                    # print(df)
                    try:
                        # print(reaction_name)
                        # print(sheet_dict[reaction_name])
                        # print(total_dic['OER'].shape)
                        df["DOI"] = df["DOI"].str.strip()
                        total_dic[sheet_dict[reaction_name]] = pd.concat([total_dic[sheet_dict[reaction_name]], df],
                                                                         ignore_index=True)
                        if check_and_update_time():
                            write_total_pickle_upload(total_dic, thermo_flag, photo_flag)
                        else:
                            st.error('System busy, upload has failed, please try again after 10 seconds!')
                        # st.write(total_dic[sheet_dict[reaction_name]])
                        return True
                    except Exception as e:
                        print(e)
                        return False


                col1, col2 = st.columns([2, 1])
                with col1:
                    st.markdown("### Upload by Excel")
                    uploaded_excel = st.file_uploader(label="Please upload the file according to the template",
                                                      type=['xlsx'])
                with col2:
                    st.markdown("### Download template here")
                    st.caption("please choose the reaction type first")
                    # 此处的文件路径可能需要和GoogleDrive链接
                    if photo_flag:
                        with open("./ExcelTemplates/photo_template.xlsx", "rb") as f:
                            excel_data = f.read()
                            template_filename = "photo_template.xlsx"
                    elif thermo_flag:
                        with open("./ExcelTemplates/thermo_template.xlsx", "rb") as f:
                            excel_data = f.read()
                            template_filename = "thermo_template.xlsx"
                    else:
                        with open("./ExcelTemplates/elec_template.xlsx", "rb") as f:
                            excel_data = f.read()
                            template_filename = "elec_template.xlsx"
                    st.download_button("Click to download",
                                       data=excel_data,
                                       file_name=template_filename,
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                # Upload Excel file
                if uploaded_excel:
                    excel_file = pd.ExcelFile(uploaded_excel)
                    sheetnames = excel_file.sheet_names
                    # 判断反应名是否存在
                    for sheetname in sheetnames:
                        # 此处为判断excel的sheetname是否能匹配上reaction name
                        if sheetname not in expected_reactions:
                            st.markdown(
                                f"<span style='color:red;'>Reaction name '{sheetname}' mismatch the catalysis type.</span>",
                                unsafe_allow_html=True)
                            continue
                        # 反应名对应的列名
                        cur_df = pd.read_excel(uploaded_excel, sheet_name=sheetname, dtype=str)
                        # 调用函数，返回错误列表
                        expected_columns = expected_reactions.get(sheetname, {}).keys()
                        errors, styles, doi_ls, df_to_upload = collect_errors_and_styles(cur_df, expected_columns,
                                                                                         sheetname, doi_database)
                        if not df_to_upload.empty:
                            with st.container(border=True):
                                col1, col2 = st.columns(2)
                                with col1:
                                    # 展示df并附带高亮
                                    st.markdown(f"#### {sheetname}: ")
                                    st.dataframe(df_to_upload.style.apply(lambda _: styles, axis=None))
                                    # Display errors if any
                                    if errors:
                                        st.markdown("#### Errors found:")
                                        for e_type in errors.keys():
                                            with st.expander(f"{e_type} errors", expanded=False):
                                                for txt in errors[e_type]:
                                                    st.write(f"- {txt}")
                                    else:
                                        if st.button(label="submit", key=f"inner_submit_{sheetname}"):
                                            if upload_cur_df(df_to_upload, sheetname):
                                                st.success("Uploaded successfully!")
                                                st.rerun()
                                            else:
                                                st.error("Reaction type mismatched ")

                                with col2:
                                    # 标题
                                    st.markdown("#### Upload your DOI here")
                                    # 使用 st.session_state 缓存 df_doi（此处为doi检查表功能，尚不完善）
                                    # if f"df_doi_{sheetname}" not in st.session_state:
                                    #     # 初始化时将 IsUploaded 设置为 False
                                    #     st.session_state[f"df_doi_{sheetname}"] = pd.DataFrame({
                                    #         "DOIs": doi_ls,
                                    #         "IsUploaded": False,  # 初始状态为未上传
                                    #     })
                                    # 获取缓存的 df_doi（此处为doi检查表功能，尚不完善）
                                    # df_doi = st.session_state[f"df_doi_{sheetname}"]
                                    # 创建表单
                                    with st.form(key=f"upload_form_{sheetname}"):
                                        selected_doi = st.selectbox("Select a DOI to Upload", doi_ls)
                                        uploaded_doi = st.file_uploader("Upload your PDF file", type=['pdf'])
                                        si_flag = st.checkbox("Please tick it if uploading the Supplementary Information(SI)")
                                        submitted = st.form_submit_button("Submit")
                                    # 如果表单提交了并且上传了文件
                                    if submitted:
                                        if uploaded_doi is not None and selected_doi is not None:
                                            # 使用 selectbox 中选择的 DOI 名称作为文件名
                                            # new_file_name = f"{selected_doi}.pdf"
                                            drive_client = init_drive_client()
                                            if si_flag:
                                                uploaded_doi.name = f"{selected_doi}_si.pdf"
                                            else:
                                                uploaded_doi.name = f"{selected_doi}.pdf"
                                            if thermo_flag:
                                                utils.save_files(uploaded_doi, "./pdfs_thermocatalysis")
                                                st_upload_file_to_drive(uploaded_doi, dir_dict["pdfs_thermocatalysis"],
                                                                        drive_client)
                                            elif photo_flag:
                                                utils.save_files(uploaded_doi, "./pdfs_photocatalysis")
                                                st_upload_file_to_drive(uploaded_doi, dir_dict["pdfs_photocatalysis"],
                                                                        drive_client)
                                            else:
                                                utils.save_files(uploaded_doi, "./pdfs")
                                                st_upload_file_to_drive(uploaded_doi, dir_dict["pdfs"], drive_client)
                                            st.success('File has been uploaded!')
                                            # st.success("File renamed and uploaded successfully!")
                                        else:
                                            st.error("No file was uploaded. Please upload a PDF file.")
                                    # 此处为doi检查表格功能，尚不完善
                                    # st.write("Check Updated DOI here:")
                                    # st.dataframe(df_doi, use_container_width=True)
                        elif reaction_name == sheetname:
                            st.markdown(
                                f"<span style='color:blue;'>After removing duplicate DOIs, reaction {sheetname} has no content.</span>",
                                unsafe_allow_html=True)

                else:
                    st.write("Please upload an Excel file to see the data.")

            st.divider()

            # 上传结构数据
            with st.container():
                # st.write(experiment_dois)
                st.markdown("### Upload Structure")
                select_upload_type = st.radio(
                    "Choose one:",
                    ('Experimental', 'Computational')
                )
                if select_upload_type == "Experimental":
                    experimental_data_path = "./computational_data/experimental_data.pkl"
                    if not os.path.exists(experimental_data_path):
                        experimental_data_path = file_from_gdrive(dir_dict["computational_data"],
                                                                  "experimental_data.pkl", init_drive_client(),
                                                                  experimental_data_path)
                        raise f"File not found {experimental_data_path}"
                    with open(experimental_data_path, "rb") as f:
                        total_exp_dic = pickle.load(f)
                        exp_df = total_exp_dic[reaction_name]
                    experiment_dois = exp_df["DOI"].unique()  # 所有
                    with st.container(border=True):
                        doi_in = st.text_input("Please enter your doi", value="10.xxxx/xxxx").strip()
                        st.divider()

                        if not is_valid_doi(doi_in):
                            st.error("Please enter a valid DOI.")
                        else:
                            if doi_in in doi_database:
                                with st.container():
                                    _col1, _col2 = st.columns(2)
                                with _col1:
                                    select_formula = st.selectbox(label="Choose a Formula",
                                                                  options=current_df[current_df["DOI"] == doi_in][
                                                                      "Formula"].unique())
                                err = uploadStruc.is_doi_name_match(doi_in, name, select_upload_type, reaction_name)
                                if doi_in in experiment_dois:
                                    if err is not None:
                                        st.error(err)
                                    else:
                                        with _col1:
                                            st.dataframe(
                                                exp_df[(exp_df["Uploader"] == name) & (exp_df["DOI"] == doi_in)])
                                if doi_in not in experiment_dois or err is None:
                                    with _col2:
                                        select_file_type = st.selectbox(label="Choose a file type",
                                                                        options=["CONTCAR", "CIF", "XYZ"])
                                        formula_struc_file = st.file_uploader(
                                            label="Upload your structure file (contcar, cif, xyz)",
                                            type=None)
                                        submit_button = st.button(
                                            f"Submit substrate Data and {select_file_type} File", )
                                        if submit_button:
                                            err = uploadStruc.upload_struc_file(select_file_type, doi_in,
                                                                                formula_struc_file, select_formula)
                                            if err is None:
                                                df = uploadStruc.upload_struc_data(exp_df,
                                                                                   select_formula,
                                                                                   doi_in,
                                                                                   name)
                                                total_exp_dic[reaction_name] = df
                                                with open("./computational_data/experimental_data.pkl", "wb") as f:
                                                    pickle.dump(total_exp_dic, f)
                                                    upload_or_replace_file("experimental_data.pkl",
                                                                           "./computational_data/experimental_data.pkl",
                                                                           "application/octet-stream",
                                                                           dir_dict["computational_data"])
                                                st.success("Uploaded successfully!")
                                            else:
                                                st.error(err)
                                    with st.container():  # 吸附物部分
                                        adsorbates = adsorbate_dic.get(reaction_name, [])
                                        # with st.form(key=f"{select_formula}_ads_file_upload"):
                                        _col1, _col2, _col3 = st.columns([3, 3, 4])
                                        with _col1:
                                            if not adsorbates:
                                                select_adsorbate = st.text_input("Please enter your adsorbate", )
                                            else:
                                                select_adsorbate = st.selectbox(label="Choose an adsorbate",
                                                                                options=adsorbates)
                                        with _col2:
                                            energy_ads = st.number_input(label="Input energy of adsorbate", )
                                            select_file_type = st.selectbox(label="Choose a file type",
                                                                            options=["CONTCAR", "CIF", "XYZ"],
                                                                            key=f"ads_file_type")
                                        with _col3:
                                            adsorbate_file = st.file_uploader(
                                                label="Upload your adsorbate_dic file (contcar, cif, xyz)",
                                                type=None)
                                        with _col1:
                                            # submit_button = st.form_submit_button(
                                            #     f"Submit substrate with {select_adsorbate} Data and {select_file_type} file", )
                                            submit_button = st.button(
                                                f"Submit substrate with {select_adsorbate} Data and {select_file_type} file", )
                                        if submit_button:
                                            # 上传文件
                                            err = uploadStruc.upload_struc_file(select_file_type, doi_in,
                                                                                adsorbate_file, select_formula,
                                                                                select_adsorbate)
                                            if err is None:
                                                # 上传数据
                                                df = uploadStruc.upload_struc_data(exp_df,
                                                                                   select_formula,
                                                                                   doi_in,
                                                                                   name,
                                                                                   select_adsorbate,
                                                                                   energy_ads)
                                                total_exp_dic[reaction_name] = df
                                                with open("./computational_data/experimental_data.pkl", "wb") as f:
                                                    pickle.dump(total_exp_dic, f)
                                                    upload_or_replace_file("experimental_data.pkl",
                                                                           "./computational_data/experimental_data.pkl",
                                                                           "application/octet-stream",
                                                                           dir_dict["computational_data"])
                                                st.success("Uploaded successfully!")
                                            else:
                                                st.error(err)
                                    st.divider()
                                    with st.container():  # 输入文件部分
                                        st.markdown("#### Upload INCAR OR KPOINTS File here")
                                        incar_file = st.file_uploader("Upload your INCAR/KPOINTS file", )
                                        incar_submit_button = st.button(label="Submit",
                                                                        key=f"{select_formula}_IK_file_upload", )
                                        if incar_submit_button and incar_file is not None:
                                            err = uploadStruc.upload_INCAR_KPOINTS_file(doi_in, incar_file,
                                                                                        select_formula)
                                            if err is None:
                                                st.success("Uploaded successfully!")
                                            else:
                                                st.error(err)
                            else:
                                st.error("No Experimental Data Exist! Please Upload Experimental Data Above")

                elif select_upload_type == "Computational":
                    computational_data_path = "./computational_data/computational_data.pkl"
                    if not os.path.exists(computational_data_path):
                        computational_data_path = file_from_gdrive(dir_dict["computational_data"],
                                                                   "computational_data.pkl", init_drive_client(),
                                                                   computational_data_path)
                        raise f"File not found {computational_data_path}"
                    with open(computational_data_path, "rb") as f:
                        total_computational_dic = pickle.load(f)
                        computational_df = total_computational_dic[reaction_name]
                    computational_dois = computational_df["DOI"].unique()
                    with st.container(border=True): # 输入DOI
                        doi_in = st.text_input("Please enter your doi", value="10.xxxx/xxxx").strip()
                        st.divider()

                        if not is_valid_doi(doi_in):
                            st.error("Please enter a valid DOI.")
                        else:
                            with st.container():  # 上传化学式对应的结构文件
                                _col1, _col2 = st.columns(2)
                            with _col1:
                                input_formula = st.text_input(label="Input a Formula", ).strip()
                            err = uploadStruc.is_doi_name_match(doi_in, name, select_upload_type, reaction_name)
                            if doi_in in computational_dois:
                                if err is not None:
                                    st.error(err)
                                else:
                                    with _col1:
                                        st.dataframe(computational_df[(computational_df["Uploader"] == name) & (computational_df["DOI"] == doi_in)])
                            if doi_in not in computational_dois or err is None:
                                if input_formula:
                                    with _col2:
                                        select_file_type = st.selectbox(label="Choose a file type",
                                                                        options=["CONTCAR", "CIF", "XYZ"],
                                                                        key=f"{input_formula}_file_type")
                                        formula_struc_file = st.file_uploader(
                                            label="Upload your structure file (contcar, cif, xyz)",
                                            type=None)
                                        submit_button = st.button(f"Submit substrate Data and {select_file_type} file",
                                                                  key=f"{input_formula}_submit_button")
                                    if submit_button:
                                        err = uploadStruc.upload_struc_file(select_file_type, doi_in, formula_struc_file, input_formula)
                                        if err is None:
                                            df = uploadStruc.upload_struc_data(computational_df, input_formula, doi_in, name)
                                            total_computational_dic[reaction_name] = df
                                            with open("./computational_data/computational_data.pkl", "wb") as f:
                                                pickle.dump(total_computational_dic, f)
                                                upload_or_replace_file("computational_data.pkl",
                                                                       "./computational_data/computational_data.pkl",
                                                                       "application/octet-stream",
                                                                       dir_dict["computational_data"])
                                            st.success("Uploaded successfully!")
                                        else:
                                            st.error(err)
                                    with st.container(): # 吸附物数据
                                        adsorbates = adsorbate_dic.get(reaction_name, [])
                                        _col1, _col2, _col3 = st.columns([3, 3, 4])
                                        with _col1:
                                            if not adsorbates:
                                                select_adsorbate = st.text_input("Please input adsorbate", key=f"{reaction_name}_ads")
                                            else:
                                                select_adsorbate = st.selectbox(label="Choose an adsorbate",
                                                                                options=adsorbates)
                                        with _col2:
                                            energy_ads = st.number_input(label="Input energy of adsorbate", )
                                            select_file_type = st.selectbox(label="Choose a file type",
                                                                            options=["CONTCAR", "CIF", "XYZ"],
                                                                            key=f"{select_adsorbate}_file_type")
                                        with _col3:
                                            adsorbate_file = st.file_uploader(
                                                label="Upload your adsorbate_dic file (contcar, cif, xyz)",
                                                type=None)
                                        with _col1:
                                            submit_button = st.button(
                                                f"Submit substrate with {select_adsorbate} Data and {select_file_type} file", )
                                        if submit_button:
                                            err = uploadStruc.upload_struc_file(select_file_type, doi_in,
                                                                                adsorbate_file, input_formula,
                                                                                select_adsorbate)
                                            if err is None:
                                                df = uploadStruc.upload_struc_data(computational_df, input_formula,
                                                                                   doi_in, name, select_adsorbate,
                                                                                   energy_ads)
                                                total_computational_dic[reaction_name] = df
                                                with open("./computational_data/computational_data.pkl", "wb") as f:
                                                    pickle.dump(total_computational_dic, f)
                                                    upload_or_replace_file("computational_data.pkl",
                                                                           "./computational_data/computational_data.pkl",
                                                                           "application/octet-stream",
                                                                           dir_dict["computational_data"])
                                                st.success("Uploaded successfully!")
                                            else:
                                                st.error(err)
                                    st.divider()
                                    with st.container(): # 输入文件上传
                                        st.markdown("#### Upload INCAR OR KPOINTS File here")
                                        incar_file = st.file_uploader("Upload your INCAR/KPOINTS file", )
                                        incar_submit_button = st.button(label="Submit", key=f"{input_formula}_IK_file_upload",)
                                        if incar_submit_button and incar_file is not None:
                                            err = uploadStruc.upload_INCAR_KPOINTS_file(doi_in, incar_file, input_formula,)
                                            if err is None:
                                                st.success("Uploaded successfully!")

                                            else:
                                                st.error(err)
                                else:
                                    st.error("Please enter a Formula first")

                    # with st.container():
                    #     st.markdown("### Upload Computational Data by ZIP")
                    #     select_upload_type = st.radio(
                    #         "Choose one:",
                    #         ('Computational Structures(without adsorption free energies)',
                    #          'Computational Structures(including adsorption free energies)')
                    #     )
                    #     # TODO 考虑是否添加到Google Drive
                    #     if select_upload_type == "Computational Structures(including adsorption free energies)":
                    #         zip_template_path = "./computational_data/template.zip"
                    #     elif select_upload_type == "Computational Structures(without adsorption free energies)":
                    #         zip_template_path = "./computational_data/template.zip"
                    #     else:
                    #         zip_template_path = None
                    #     with open(zip_template_path, "rb") as zip_template_file:
                    #         zip_data = zip_template_file.read()
                    #     if zip_data:
                    #         # 创建下载按钮
                    #         st.download_button(
                    #             label="Download ZIP Template",
                    #             data=zip_data,
                    #             file_name="template.zip",
                    #             mime="application/zip"
                    #         )
                    #     else:
                    #         st.error("ZIP file not found.")
                    #     zip_file = st.file_uploader("Upload ZIP here", "zip")
                    #     _col1, _col2 = st.columns(2)
                    #     if zip_file is not None:
                    #         file_validator = uploadStrucZipClass.FileValidator(zip_file, select_upload_type, name, reaction_name)
                    #         with _col1:
                    #             __col1, __col2 = st.columns(2)
                    #         with _col2:
                    #             submit_button = st.button("Submit")
                    #         with __col1:
                    #             check_toggle = st.toggle("Check and Show Your Data")
                    #         if check_toggle:
                    #             expect_df, ved, err = file_validator.check_all()
                    #             with _col1:
                    #                 st.markdown("##### Expect Data")
                    #                 st.dataframe(ved)
                    #             with _col2:
                    #                 st.markdown("##### Got Data")
                    #                 st.dataframe(pd.DataFrame(file_validator.valid_data))
                    #             with st.expander("ERROR"):
                    #                 for e in file_validator.error_messages:
                    #                     st.write(e)
                    #         if submit_button:
                    #             if check_toggle and submit_button and file_validator.error_messages is None:
                    #                 err = file_validator.submit_data2DB()
                    #                 if err is None:
                    #                     uploadStrucbyZIP.upload_files_to_google_drive(file_validator.checked_files, dir_dict)
                    #                     st.success("Uploaded successfully!")
                    #             else:
                    #                 st.warning("Please Check your data first")









                            # if select_upload_type == "Computational Structures(including adsorption free energies)":
                            #     with _col1:
                            #         __col1, __col2 = st.columns(2)
                            #     with __col1:
                            #         check_toggle = st.toggle("Check and Show Your Data")
                    #             if check_toggle:
                    #                 # 调用后端函数处理
                    #                 err, invalid_doi, unmatched_doi, valid_entries, checked_files, infos = uploadStrucbyZIP.handle_uploaded_file(
                    #                     zip_file, select_upload_type, name, reaction_name)
                    #                 if err:
                    #                     with _col1:
                    #                         st.error(err)
                    #                 else:
                    #                     with _col1:
                    #                         for err_name in infos.keys():
                    #                             with st.expander(f"{err_name.upper()}_ERR", expanded=True):
                    #                                 if infos[err_name]:
                    #                                     st.write(infos[err_name])
                    #                     if not invalid_doi.empty:
                    #                         with _col1:
                    #                             st.error(f"You uploaded Invalid DOI: {invalid_doi['DOI'].unique()}")
                    #                     if not unmatched_doi.empty:
                    #                         with _col1:
                    #                             st.error(
                    #                                 f"{unmatched_doi['DOI'].unique()} has been uploaded by others")
                    #                     if not valid_entries.empty:
                    #                         # 展示合法的上传数据
                    #                         st.dataframe(valid_entries)
                    #                         with __col2:
                    #                             # submit_button = st.button(label="Submit Data",
                    #                             #                           disabled=(infos is not None))
                    #                             submit_button = st.button(label="Submit Data")
                    #                         if submit_button:
                    #                             err = uploadStrucbyZIP.submit_data_to_file(valid_entries, reaction_name)
                    #                             if err is None:
                    #                                 uploadStrucbyZIP.upload_files_to_google_drive(checked_files, dir_dict)
                    #                                 st.success("Uploaded Data successfully!")
                    #                             else:
                    #                                 st.error(err)
                    #         elif select_upload_type == "Computational Structures(without adsorption free energies)":
                    #             _, _, _, df, checked_files, infos = uploadStrucbyZIP.handle_uploaded_file(zip_file, select_upload_type, name, reaction_name)
                    #             with _col1:
                    #                 if infos is not None:
                    #                     for err_name in infos.keys():
                    #                         with st.expander(f"{err_name.upper()}_ERR", expanded=True):
                    #                             if infos[err_name]:
                    #                                 st.write(infos[err_name])
                    #                 __col1, __col2 = st.columns(2)
                    #             with __col1:
                    #                 st.dataframe(df)
                    #             with __col2:
                    #                 # submit_button = st.button(label="Submit Data", disabled=(infos is not None))
                    #                 submit_button = st.button(label="Submit Data")
                    #             if submit_button:
                    #                 err = uploadStrucbyZIP.submit_data_to_file(df, reaction_name)
                    #                 if err is None:
                    #                     uploadStrucbyZIP.upload_files_to_google_drive(checked_files, dir_dict)
                    #                     st.success("Uploaded Data successfully!")
                    #                 else:
                    #                     st.error(err)
                    #     else:
                    #         st.error("Please upload a zip file.")













